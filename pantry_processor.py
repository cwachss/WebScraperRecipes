# here we shall read in the pantry. I haven't decided if I should double store it or not.
import openpyxl as xl

# not sure how to use this- seems a bit difficult. I think we shall have to make the human do substitutions.
ingredient_substitutes = {
    "butter": "margarine",
    "vanilla sugar": "vanilla extract",
    "almond milk": "milk",
    "chocolate chunks": "chocolate chips",
    "confectioner's sugar": "powdered sugar",
    "icing sugar": "powdered sugar"
}


def read_pantry(pantry, pantry_number):
    pantry_workbook = xl.load_workbook(pantry)
    pantry_sheet = pantry_workbook[pantry_number]
    ingredients = []
    for row in range(2, pantry_sheet.max_row + 1):
        ingredient = pantry_sheet.cell(row, 1).value
        ingredients.append(ingredient)

    return ingredients


# test
# print(read_pantry('Pantry.xlsx', 'Pantry1'))
